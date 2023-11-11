from flask import Flask, request, send_file, render_template, redirect, url_for
import os
import pandas as pd
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from xml.dom import minidom

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '/tmp/'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx'}

@app.route('/', methods=['GET'])
def index():
    # Render the HTML page for file upload
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)

        file = request.files['file']
        group_name = request.form['group_name']
        column_name = request.form['column_name']
        match_value = request.form['match_value']

        if file.filename == '' or not allowed_file(file.filename):
            return redirect(request.url)

        filename = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filename)

        # Open the workbook and the first sheet
        wb = load_workbook(filename=filename)
        sheet = wb.active

        # Create a DataFrame from the Excel file
        df = pd.read_excel(filename, engine='openpyxl')

        # Check if the column exists
        if column_name not in df.columns:
            return f"The column '{column_name}' does not exist in the Excel file.", 400

        # Look for hyperlinks in the specified column and replace the text with the actual URL
        if 'Secret Server URL' in df.columns:
            for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
                cell = row[df.columns.get_loc('Secret Server URL')]
                if cell.hyperlink:
                    df.at[cell.row - 1, 'Secret Server URL'] = cell.hyperlink.target

        matched_df = df[df[column_name].astype(str).str.strip() == match_value.strip()]

        rdg_content = generate_rdg(matched_df, group_name)
        processed_filename = f"processed_{file.filename.rsplit('.', 1)[0]}.rdg"
        processed_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

        with open(processed_filepath, 'w') as rdg_file:
            rdg_file.write(rdg_content)

        return redirect(url_for('download_file', filename=processed_filename))

    return 'File upload error'

def prettify_xml(element):
    """Return a pretty-printed XML string for the Element."""
    rough_string = ET.tostring(element, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ")

def generate_rdg(df, group_name):
    root = ET.Element('RDCMan', programVersion="2.93", schemaVersion="3")
    file_node = ET.SubElement(root, 'file')
    group_node = ET.SubElement(file_node, 'group')
    properties_node = ET.SubElement(group_node, 'properties')
    ET.SubElement(properties_node, 'expanded').text = 'True'
    ET.SubElement(properties_node, 'name').text = group_name

    for index, row in df.iterrows():
        server_node = ET.SubElement(group_node, 'server')
        properties_node = ET.SubElement(server_node, 'properties')
        ET.SubElement(properties_node, 'displayName').text = str(row['FQDN'])
        ET.SubElement(properties_node, 'name').text = str(row['IP Address'])
        comment_text = f"Configuration Item :{row['Configuration Item Name']}"

        if 'Secret Server URL' in df.columns and pd.notnull(row['Secret Server URL']):
            secret_url = row['Secret Server URL']
        else:
            secret_url = "URL Not Available"

        comment_text += f"\nSS URL:{secret_url}"
        ET.SubElement(properties_node, 'comment').text = comment_text

    return prettify_xml(root)

@app.route('/downloads/<filename>', methods=['GET'])
def download_file(filename):
    download_folder = app.config['UPLOAD_FOLDER']
    file_path = os.path.join(download_folder, filename)
    
    if not os.path.isfile(file_path):
        return "File not found.", 404

    response = send_file(file_path, as_attachment=True, download_name=filename)
    os.remove(file_path)
    
    return response

if __name__ == '__main__':
    app.run(debug=True)
